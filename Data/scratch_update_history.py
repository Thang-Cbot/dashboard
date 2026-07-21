import pandas as pd
import openpyxl
import os

EXCEL_FILE = os.path.join("USDA Reports", "CBOT_Reports_History.xlsx")

def clean_and_backfill():
    wb = openpyxl.load_workbook(EXCEL_FILE)
    
    # 1. Clean empty rows
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        # Start from bottom and delete rows where Column A is None
        for row in range(sheet.max_row, 1, -1):
            val = sheet.cell(row=row, column=1).value
            if val is None or str(val).strip() == "" or str(val).lower() == "nan":
                sheet.delete_rows(row, 1)

    # 2. Backfill missing previous period if not exist
    
    # Export Sales
    sales_ws = wb["Export_Sales"]
    has_june26 = False
    for r in range(2, sales_ws.max_row + 1):
        if "June 26" in str(sales_ws.cell(r, 1).value):
            has_june26 = True
            break
            
    if not has_june26:
        # We need to insert it BEFORE the July 3-9 row.
        # Find July 3-9 row
        jul3_row = None
        for r in range(2, sales_ws.max_row + 1):
            if "July 3-9" in str(sales_ws.cell(r, 1).value):
                jul3_row = r
                break
        
        insert_idx = jul3_row if jul3_row else sales_ws.max_row + 1
        sales_ws.insert_rows(insert_idx)
        # 'Ngày Báo Cáo', 'ZW Net Sales', 'ZW Thay đổi (%)', 'ZW Ship', 'ZW Out', ' ', 'ZC Net Sales', 'ZC Thay đổi (%)', 'ZC Ship', 'ZC Out'
        sales_ws.cell(insert_idx, 1).value = "June 26-July 2, 2026"
        sales_ws.cell(insert_idx, 2).value = 313.1
        sales_ws.cell(insert_idx, 3).value = 0 # unknown
        sales_ws.cell(insert_idx, 4).value = 0 # unknown
        sales_ws.cell(insert_idx, 7).value = 565.8
        sales_ws.cell(insert_idx, 8).value = 0 # unknown
        sales_ws.cell(insert_idx, 9).value = 0 # unknown

    # Export Inspections
    insp_ws = wb["Export_Inspections"]
    has_jul09 = False
    for r in range(2, insp_ws.max_row + 1):
        if "JUL 09" in str(insp_ws.cell(r, 1).value):
            has_jul09 = True
            break
            
    if not has_jul09:
        jul16_row = None
        for r in range(2, insp_ws.max_row + 1):
            if "JUL 16" in str(insp_ws.cell(r, 1).value):
                jul16_row = r
                break
        
        insert_idx = jul16_row if jul16_row else insp_ws.max_row + 1
        insp_ws.insert_rows(insert_idx)
        # 'Ngày Báo Cáo', 'ZW Giao Hàng', 'ZW Thay đổi (%)', ' ', 'ZC Giao Hàng', 'ZC Thay đổi (%)'
        insp_ws.cell(insert_idx, 1).value = "JUL 09, 2026"
        insp_ws.cell(insert_idx, 2).value = 396.3
        insp_ws.cell(insert_idx, 3).value = 0
        insp_ws.cell(insert_idx, 5).value = 1554.6
        insp_ws.cell(insert_idx, 6).value = 0

    # Crop Progress
    cp_ws = wb["Crop_Progress"]
    has_jul13 = False
    for r in range(2, cp_ws.max_row + 1):
        if "07-13" in str(cp_ws.cell(r, 1).value) or "13/07" in str(cp_ws.cell(r, 1).value):
            has_jul13 = True
            break
            
    if not has_jul13:
        jul20_row = None
        for r in range(2, cp_ws.max_row + 1):
            if "07-20" in str(cp_ws.cell(r, 1).value) or "20/07" in str(cp_ws.cell(r, 1).value):
                jul20_row = r
                break
                
        insert_idx = jul20_row if jul20_row else cp_ws.max_row + 1
        cp_ws.insert_rows(insert_idx)
        cp_ws.cell(insert_idx, 1).value = "13/07/2026"
        cp_ws.cell(insert_idx, 3).value = "60% thu hoạch"
        cp_ws.cell(insert_idx, 7).value = "65% G/E"
        cp_ws.cell(insert_idx, 9).value = "97% đã gieo trồng"
        cp_ws.cell(insert_idx, 11).value = "68% Good to Excellent"

    wb.save(EXCEL_FILE)
    print("Done cleaning and backfilling.")

if __name__ == "__main__":
    clean_and_backfill()
