import json
import os
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def extract_number(s):
    if s is None: return None
    if isinstance(s, (int, float)): return s
    s = str(s).lower().replace(',', '')
    import re
    m = re.search(r'[-+]?\d*\.\d+|\d+', s)
    if m: return float(m.group())
    return None

def create_or_load_workbook(path):
    if path.exists():
        return openpyxl.load_workbook(path)
    return openpyxl.Workbook()

def setup_sheet(wb, sheet_name, headers, col_widths):
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.create_sheet(sheet_name)
    
    if ws.max_row == 1 and ws.cell(row=1, column=1).value is None:
        header_fill = PatternFill(start_color='1F497D', end_color='1F497D', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)
        align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = h
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = align
            cell.border = border
            if h == ' ':
                ws.column_dimensions[get_column_letter(col)].width = 2
            else:
                ws.column_dimensions[get_column_letter(col)].width = col_widths.get(col, 15)
        ws.freeze_panes = 'A2'
    return ws

def append_row(ws, row_data, date_val):
    for r in range(2, ws.max_row + 1):
        if ws.cell(row=r, column=1).value == date_val:
            return # Skip if date exists
    
    ws.append(row_data)
    row_idx = ws.max_row
    for col in range(2, len(row_data) + 1):
        cell = ws.cell(row=row_idx, column=col)
        if isinstance(cell.value, (int, float)):
            header = ws.cell(row=1, column=col).value
            if header and '%' in header:
                cell.number_format = '0.00'
            else:
                cell.number_format = '#,##0'

def update_excel_history(data_dir=Path('../output'), excel_path=Path('../../USDA Reports/CBOT_Reports_History.xlsx')):
    fund_path = data_dir / 'fundamental_data.json'
    acreage_path = data_dir / 'acreage_data.json'
    
    def get_json(p):
        if p.exists():
            with open(p, 'r', encoding='utf-8') as f:
                return json.load(f)
        return {}
    
    fund = get_json(fund_path)
    acreage = get_json(acreage_path)
    wb = create_or_load_workbook(excel_path)
    
    # 1. Export Sales
    sales_headers = ['Ngày Báo Cáo', 'ZW Net Sales (tấn)', 'ZW Thay đổi (%)', 'ZW Shipments (tấn)', 'ZW Outstanding (tấn)', ' ', 'ZC Net Sales (tấn)', 'ZC Thay đổi (%)', 'ZC Shipments (tấn)', 'ZC Outstanding (tấn)']
    ws_sales = setup_sheet(wb, 'Export_Sales', sales_headers, {1:15, 2:18, 3:15, 4:18, 5:20, 7:18, 8:15, 9:18, 10:20})
    zc_sales = fund.get('ZC', {}).get('export_sales_weekly', {})
    zw_sales = fund.get('ZW', {}).get('export_sales_weekly', {})
    if zc_sales and zw_sales:
        curr_date = zc_sales.get('latest_date')
        if curr_date:
            append_row(ws_sales, [
                curr_date,
                extract_number(zw_sales.get('latest_net_sales')), extract_number(zw_sales.get('pct_change')),
                extract_number(zw_sales.get('latest_shipments')), extract_number(zw_sales.get('outstanding_sales')),
                '',
                extract_number(zc_sales.get('latest_net_sales')), extract_number(zc_sales.get('pct_change')),
                extract_number(zc_sales.get('latest_shipments')), extract_number(zc_sales.get('outstanding_sales'))
            ], curr_date)

    # 2. Export Inspections
    insp_headers = ['Ngày Báo Cáo', 'ZW Giao Hàng (tấn)', 'ZW Thay đổi (%)', ' ', 'ZC Giao Hàng (tấn)', 'ZC Thay đổi (%)']
    ws_insp = setup_sheet(wb, 'Export_Inspections', insp_headers, {1:25, 2:20, 3:15, 5:20, 6:15})
    zc_insp = fund.get('ZC', {}).get('exports', {})
    zw_insp = fund.get('ZW', {}).get('exports', {})
    if zc_insp and zw_insp:
        curr_date = zc_insp.get('latest_date')
        if curr_date:
            append_row(ws_insp, [
                curr_date,
                extract_number(zw_insp.get('latest')), extract_number(zw_insp.get('pct_change')),
                '',
                extract_number(zc_insp.get('latest')), extract_number(zc_insp.get('pct_change'))
            ], curr_date)

    # 3. Crop Progress
    cp_headers = ['Ngày Báo Cáo', 'ZW Winter Gieo (%)', 'ZW Winter Thu Hoạch (%)', 'ZW Winter G/E (%)', 'ZW Spring Gieo (%)', 'ZW Spring Thu Hoạch (%)', 'ZW Spring G/E (%)', ' ', 'ZC Gieo Trồng (%)', 'ZC Thu Hoạch (%)', 'ZC G/E (%)']
    ws_cp = setup_sheet(wb, 'Crop_Progress', cp_headers, {1:15})
    zc_cp_p = fund.get('ZC', {}).get('us_planting', {})
    if zc_cp_p:
        curr_date = zc_cp_p.get('latest_date')
        if curr_date:
            append_row(ws_cp, [
                curr_date,
                None, None, None, None, None, None, '',
                extract_number(zc_cp_p.get('latest')), extract_number(fund.get('ZC', {}).get('harvest_progress', {}).get('latest')), extract_number(fund.get('ZC', {}).get('crop_condition', {}).get('latest'))
            ], curr_date)

    # 4. WASDE
    w_headers = ['Kỳ Báo Cáo', 'ZW Tồn Kho Mỹ (tr bu)', 'ZW Tồn Kho TG (tr tấn)', ' ', 'ZC Tồn Kho Mỹ (tr bu)', 'ZC Tồn Kho TG (tr tấn)']
    ws_w = setup_sheet(wb, 'WASDE', w_headers, {1:15, 2:25, 3:25, 5:25, 6:25})
    zc_w_us = fund.get('ZC', {}).get('us_ending_stocks', {})
    if zc_w_us:
        curr_date = zc_w_us.get('latest_date', 'Tháng 6/2026')
        append_row(ws_w, [
            curr_date,
            None, None, '',
            extract_number(zc_w_us.get('current')), extract_number(fund.get('ZC', {}).get('global_ending_stocks', {}).get('current'))
        ], curr_date)

    # 5. Acreage
    ac_headers = ['Kỳ Báo Cáo', 'ZW All Diện Tích (Tr mẫu)', 'ZW Thay đổi (%)', 'ZW Winter (Tr mẫu)', 'ZW Spring (Tr mẫu)', ' ', 'ZC Ngô (Tr mẫu)', 'ZC Thay đổi (%)']
    ws_ac = setup_sheet(wb, 'Acreage', ac_headers, {1:15, 2:25, 3:15, 4:25, 5:25, 7:20, 8:15})
    if 'commodities' in acreage:
        curr_date = 'Tháng 6/2026'
        zw_ac = acreage['commodities'].get('ZW', {})
        zc_ac = acreage['commodities'].get('ZC', {})
        append_row(ws_ac, [
            curr_date,
            extract_number(zw_ac.get('latest_planted_mln_acres')), extract_number(zw_ac.get('yoy_pct')),
            None, None, '',
            extract_number(zc_ac.get('latest_planted_mln_acres')), extract_number(zc_ac.get('yoy_pct'))
        ], curr_date)

    if 'Sheet' in wb.sheetnames: del wb['Sheet']
    wb.save(excel_path)
    print(f'Successfully updated {excel_path}')

if __name__ == '__main__':
    script_dir = Path(__file__).parent
    update_excel_history(data_dir=script_dir.parent / 'output', excel_path=script_dir.parent.parent / 'USDA Reports' / 'CBOT_Reports_History.xlsx')
