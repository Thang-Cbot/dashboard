import os
import openpyxl

EXCEL_FILE = os.path.join(os.path.dirname(__file__), "..", "USDA Reports", "CBOT_Reports_History.xlsx")

def _get_sheet(wb, sheet_name):
    if sheet_name in wb.sheetnames:
        return wb[sheet_name]
    return None

def _is_duplicate_date(sheet, date_str):
    """Kiểm tra xem ngày (cột A) ở dòng cuối cùng có trùng với date_str không."""
    if sheet.max_row <= 1:
        return False
    last_val = sheet.cell(row=sheet.max_row, column=1).value
    return str(last_val).strip() == str(date_str).strip()

def log_export_sales(date_str, zc_sales, zc_pct, zc_ship, zc_out, zw_sales, zw_pct, zw_ship, zw_out):
    """Ghi Export Sales vào Excel."""
    if not os.path.exists(EXCEL_FILE): return
    try:
        wb = openpyxl.load_workbook(EXCEL_FILE)
        sheet = _get_sheet(wb, "Export_Sales")
        if not sheet: return
        
        if _is_duplicate_date(sheet, date_str):
            print(f"  [Excel Logger] Export_Sales kỳ '{date_str}' đã tồn tại. Bỏ qua.")
            return

        # Cột: A: Ngày, B: ZW Sales, C: ZW %, D: ZW Ship, E: ZW Out, F: trống, G: ZC Sales, H: ZC %, I: ZC Ship, J: ZC Out
        sheet.append([date_str, zw_sales, zw_pct, zw_ship, zw_out, "", zc_sales, zc_pct, zc_ship, zc_out])
        wb.save(EXCEL_FILE)
        print(f"  [Excel Logger] Đã lưu Export_Sales kỳ '{date_str}'.")
    except Exception as e:
        print(f"  [Excel Logger] Lỗi khi lưu Export_Sales: {e}")

def log_export_inspections(date_str, zc_vol, zc_pct, zw_vol, zw_pct):
    """Ghi Export Inspections vào Excel."""
    if not os.path.exists(EXCEL_FILE): return
    try:
        wb = openpyxl.load_workbook(EXCEL_FILE)
        sheet = _get_sheet(wb, "Export_Inspections")
        if not sheet: return

        if _is_duplicate_date(sheet, date_str):
            print(f"  [Excel Logger] Export_Inspections kỳ '{date_str}' đã tồn tại. Bỏ qua.")
            return

        # Cột: A: Ngày, B: ZW Vol, C: ZW %, D: trống, E: ZC Vol, F: ZC %
        sheet.append([date_str, zw_vol, zw_pct, "", zc_vol, zc_pct])
        wb.save(EXCEL_FILE)
        print(f"  [Excel Logger] Đã lưu Export_Inspections kỳ '{date_str}'.")
    except Exception as e:
        print(f"  [Excel Logger] Lỗi khi lưu Export_Inspections: {e}")

def log_wasde(date_str, zc_us, zc_world, zw_us, zw_world):
    """Ghi WASDE vào Excel."""
    if not os.path.exists(EXCEL_FILE): return
    try:
        wb = openpyxl.load_workbook(EXCEL_FILE)
        sheet = _get_sheet(wb, "WASDE")
        if not sheet: return

        if _is_duplicate_date(sheet, date_str):
            print(f"  [Excel Logger] WASDE kỳ '{date_str}' đã tồn tại. Bỏ qua.")
            return

        # Cột: A: Kỳ, B: ZW US, C: ZW TG, D: trống, E: ZC US, F: ZC TG
        sheet.append([date_str, zw_us, zw_world, "", zc_us, zc_world])
        wb.save(EXCEL_FILE)
        print(f"  [Excel Logger] Đã lưu WASDE kỳ '{date_str}'.")
    except Exception as e:
        print(f"  [Excel Logger] Lỗi khi lưu WASDE: {e}")

def log_crop_progress(date_str, 
                      zw_w_plant, zw_w_harv, zw_w_ge,
                      zw_s_plant, zw_s_harv, zw_s_ge,
                      zc_plant, zc_harv, zc_ge):
    """Ghi Crop Progress vào Excel."""
    if not os.path.exists(EXCEL_FILE): return
    try:
        wb = openpyxl.load_workbook(EXCEL_FILE)
        sheet = _get_sheet(wb, "Crop_Progress")
        if not sheet: return

        if _is_duplicate_date(sheet, date_str):
            print(f"  [Excel Logger] Crop_Progress kỳ '{date_str}' đã tồn tại. Bỏ qua.")
            return

        # Cột: A: Ngày, B->D: ZW Winter (Gieo, Thu, G/E), E->G: ZW Spring (Gieo, Thu, G/E), H: trống, I->K: ZC (Gieo, Thu, G/E)
        sheet.append([
            date_str, 
            zw_w_plant, zw_w_harv, zw_w_ge,
            zw_s_plant, zw_s_harv, zw_s_ge,
            "",
            zc_plant, zc_harv, zc_ge
        ])
        wb.save(EXCEL_FILE)
        print(f"  [Excel Logger] Đã lưu Crop_Progress kỳ '{date_str}'.")
    except Exception as e:
        print(f"  [Excel Logger] Lỗi khi lưu Crop_Progress: {e}")
